import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import math
import statistics
from scipy.optimize import curve_fit
from openpyxl import load_workbook


a51 = pd.read_csv("5.1.csv",sep=';',header=0, decimal=",")  
a52 = pd.read_csv('5.2.csv',sep=';',header=0, decimal=",")     
a53 = pd.read_csv('5.3.csv',sep=';',header=0, decimal=",")     
a61 = pd.read_csv('6.1.csv',sep=';',header=0, decimal=",")   
a62 = pd.read_csv('6.2.csv',sep=';',header=0, decimal=",")   
a63 = pd.read_csv('6.3.csv',sep=';',header=0, decimal=",")    
a71 = pd.read_csv('7.1.csv',sep=';',header=0, decimal=",")  
a72 = pd.read_csv('7.2.csv',sep=';',header=0, decimal=",")   
a73 = pd.read_csv('7.3.csv',sep=';',header=0, decimal=",")   
a81 = pd.read_csv('8.1.csv',sep=';',header=0, decimal=",")   
a82 = pd.read_csv('8.2.csv',sep=';',header=0, decimal=",")     
a83 = pd.read_csv('8.3.csv',sep=';',header=0, decimal=",")
a91 = pd.read_csv('9.1.csv',sep=';',header=0, decimal=",")
a92 = pd.read_csv('9.2.csv',sep=';',header=0, decimal=",")
a93 = pd.read_csv('9.3EDITAR.csv',sep=';',header=0, decimal=",")
a101 = pd.read_csv('10.1.csv',sep=';',header=0, decimal=",")
a102 = pd.read_csv('10.2.csv',sep=';',header=0, decimal=",")
a103 = pd.read_csv('10.3.csv',sep=';',header=0, decimal=",")
a111 = pd.read_csv('11.1.csv',sep=';',header=0, decimal=",")
a112 = pd.read_csv('11.2.csv',sep=';',header=0, decimal=",")
a113 = pd.read_csv('11.3.csv',sep=';',header=0, decimal=",")
a121 = pd.read_csv('12.1.csv',sep=';',header=0, decimal=",")
a122 = pd.read_csv('12.2.csv',sep=';',header=0, decimal=",")
a123 = pd.read_csv('12.3.csv',sep=';',header=0, decimal=",")



datos = [a51,a52,a53,a61,a62,a63,a71,a72,a73,a81,a82,a83,a91,a92,a93,a101,a102,a103,a111,a112,a113,a121,a122,a123]
for i in datos:
    i.columns =  ['tiempo','v1']  
    

    
    
columnaTiempo = a51['tiempo']    
columnaTiempo = columnaTiempo.to_frame()
    
    
a101 = a101[a101.tiempo%5==0]
a122 = a122[a122.tiempo%5==0]
a101 = a101.reset_index(drop= True)

a122= a122.reset_index(drop= True)

ph5=a51
ph5.insert(2,'v2',a52.v1)
ph5.insert(3,'v3',a53.v1)
 
#print(ph5.head())
ph5= ph5.drop(columns = 'tiempo')
ph5['promedio']= ph5.mean(axis = 1)


ph6=a61
ph6.insert(2,'v2',a62.v1)
ph6.insert(3,'v3',a63.v1)
#print(ph5.head())
ph6= ph6.drop(columns = 'tiempo')
ph6['promedio']= ph6.mean(axis = 1)
 

ph7=a71
ph7.insert(2,'v2',a72.v1)
ph7.insert(3,'v3',a73.v1)
#print(ph5.head())
ph7= ph7.drop(columns = 'tiempo')
ph7['promedio']= ph7.mean(axis = 1)
 

ph8=a81
ph8.insert(2,'v2',a82.v1)
ph8.insert(3,'v3',a83.v1)
#print(ph5.head())
ph8= ph8.drop(columns = 'tiempo')
ph8['promedio']= ph8.mean(axis = 1)
 

ph9=a91
ph9.insert(2,'v2',a92.v1)
ph9.insert(3,'v3',a93.v1)
#print(ph5.head())
ph9= ph9.drop(columns = 'tiempo')
ph9['promedio']= ph9.mean(axis = 1)
 


ph10=a101
ph10.insert(2,'v2',a102.v1)
ph10.insert(3,'v3',a103.v1)
#print(ph5.head())
ph10= ph10.drop(columns = 'tiempo')
ph10['promedio']= ph10.mean(axis = 1)
 


ph11=a111
ph11.insert(2,'v2',a112.v1)
ph11.insert(3,'v3',a113.v1)
#print(ph5.head())
ph11= ph11.drop(columns = 'tiempo')
ph11['promedio']= ph11.mean(axis = 1)
 


ph12=a121
ph12.insert(2,'v2',a122.v1)
ph12.insert(3,'v3',a123.v1)
#print(ph5.head())
ph12= ph12.drop(columns = 'tiempo')
ph12['promedio']= ph12.mean(axis = 1)
 




 


ph5.insert(4, 'tiempo', columnaTiempo.tiempo)
ph6.insert(4, 'tiempo', columnaTiempo.tiempo)
ph7.insert(4, 'tiempo', columnaTiempo.tiempo)
ph8.insert(4, 'tiempo', columnaTiempo.tiempo)
ph9.insert(4, 'tiempo', columnaTiempo.tiempo)
ph10.insert(4, 'tiempo', columnaTiempo.tiempo)
ph11.insert(4, 'tiempo', columnaTiempo.tiempo)
ph12.insert(4, 'tiempo', columnaTiempo.tiempo)


book = load_workbook('TodoPH.xlsx')



for i in range (5,13):
    cuerda = 'ph' +str(i)
    book.create_sheet(cuerda)


LOSPH= [ph5,ph6,ph7,ph8,ph9,ph10,ph11,ph12]

writer = pd.ExcelWriter('TodoPH.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}





k = 5
for i,j in zip(LOSPH, book.worksheets):
    cuerda = 'ph' + str(k)
    i.to_excel(writer,sheet_name=cuerda, startrow=writer.sheets[cuerda].max_row, index = False,header= False)
    k+=1


writer.save()





coef5 = np.polyfit(ph5.tiempo, ph5.v1, 1)
poly1d_fn = np.poly1d(coef5) 
fig, ax = plt.subplots()
ax.plot(ph5.tiempo, ph5.v1, label = 'Versuch 1')
ax.plot(ph5.tiempo, ph5.v2, label = 'Versuch 2')
ax.plot(ph5.tiempo, ph5.v3, label = 'Versuch 3')
ax.plot(ph5.tiempo, ph5.promedio, label = 'Durchschnitt')
ax.plot(ph5.tiempo, poly1d_fn(ph5.tiempo), label ='regresion')

ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.legend()
ax.set_title('ph 5')



coef6 = np.polyfit(ph6.tiempo, ph6.v2, 1)
poly1d_fn = np.poly1d(coef6) 
fig, ax = plt.subplots()
ax.plot(ph6.tiempo, ph6.v1, label = 'Versuch 1')
ax.plot(ph6.tiempo, ph6.v2, label = 'Versuch 2')
ax.plot(ph6.tiempo, ph6.v3, label = 'Versuch 3')
ax.plot(ph6.tiempo, ph6.promedio, label = 'Durchschnitt')
ax.plot(ph6.tiempo, poly1d_fn(ph6.tiempo), label ='regresion')

ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.legend()
ax.set_title('ph 6')


coef7 = np.polyfit(ph7.tiempo, ph7.v1, 1)
poly1d_fn = np.poly1d(coef7) 
fig, ax = plt.subplots()
ax.plot(ph7.tiempo, ph7.v1, label = 'Versuch 1')
ax.plot(ph7.tiempo, ph7.v2, label = 'Versuch 2')
ax.plot(ph7.tiempo, ph7.v3, label = 'Versuch 3')
ax.plot(ph7.tiempo, ph7.promedio, label = 'Durchschnitt')
ax.plot(ph7.tiempo, poly1d_fn(ph7.tiempo), label ='regresion')

ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.set_title('ph 7')
ax.legend()


coef8 = np.polyfit(ph8.tiempo, ph8.promedio, 1)
poly1d_fn = np.poly1d(coef8) 
fig, ax = plt.subplots()
ax.plot(ph8.tiempo, ph8.v1, label = 'Versuch 1')
ax.plot(ph8.tiempo, ph8.v2, label = 'Versuch 2')
ax.plot(ph8.tiempo, ph8.v3, label = 'Versuch 3')
ax.plot(ph8.tiempo, ph8.promedio, label = 'Durchschnitt')
ax.plot(ph8.tiempo, poly1d_fn(ph8.tiempo), label ='regresion')

ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.set_title('ph 8')
ax.legend()



coef9 = np.polyfit(ph9.tiempo, ph9.v2, 1)
poly1d_fn = np.poly1d(coef9) 
fig, ax = plt.subplots()
ax.plot(ph9.tiempo, ph9.v1, label = 'Versuch 1')
ax.plot(ph9.tiempo, ph9.v2, label = 'Versuch 2')
ax.plot(ph9.tiempo, ph9.v3, label = 'Versuch 3')
ax.plot(ph9.tiempo, ph9.promedio, label = 'Durchschnitt')
ax.plot(ph9.tiempo, poly1d_fn(ph9.tiempo), label ='regresion')

ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.set_title('ph 9')
ax.legend()



coef10 = np.polyfit(ph10.tiempo, ph10.v2, 1)
poly1d_fn = np.poly1d(coef10) 


fig, ax = plt.subplots()
ax.plot(ph10.tiempo, ph10.v1, label = 'Versuch 1')
ax.plot(ph10.tiempo, ph10.v2, label = 'Versuch 2')
ax.plot(ph10.tiempo, ph10.v3, label = 'Versuch 3')
ax.plot(ph10.tiempo, ph10.promedio, label = 'Durchschnitt')
ax.plot(ph10.tiempo, poly1d_fn(ph10.tiempo), label ='regresion')

ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.set_title('ph 10')
ax.legend()




coef = np.polyfit(ph11.tiempo, ph11.v3, 1)
poly1d_fn = np.poly1d(coef) 
fig, ax = plt.subplots()
ax.plot(ph11.tiempo, ph11.v1, label = 'Versuch 1')
ax.plot(ph11.tiempo, ph11.v2, label = 'Versuch 2')
ax.plot(ph11.tiempo, ph11.v3, label = 'Versuch 3')
ax.plot(ph11.tiempo, ph11.promedio, label = 'Durchschnitt')
ax.plot(ph11.tiempo, poly1d_fn(ph11.tiempo), label ='regresion')
#ax.scatter(ph11.tiempo,pred)
ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.set_title('ph 11')
ax.legend()



fig, ax = plt.subplots()
ax.plot(ph11.tiempo, ph7.v1, label = 'ph 7')
ax.plot(ph11.tiempo, ph6.promedio, label = 'ph 6')
ax.plot(ph11.tiempo, ph5.v1, label = 'ph 5')

#ax.scatter(ph11.tiempo,pred)
ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.set_title('pH Werte: 5 - 7')
ax.legend()




fig, ax = plt.subplots()
ax.plot(ph11.tiempo, ph7.v1, label = 'ph 7')
ax.plot(ph11.tiempo, ph8.v1, label = 'ph 8')
ax.plot(ph11.tiempo, ph9.v2, label = 'ph 9')
ax.plot(ph11.tiempo, ph10.v3, label = 'ph 10')
ax.plot(ph11.tiempo, ph11.v1, label = 'ph 11')
#ax.scatter(ph11.tiempo,pred)
ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.set_title('pH Werte: 7 - 11')
ax.legend()




fig, ax = plt.subplots()
ax.plot(ph11.tiempo, ph7.promedio, label = 'ph 7')
ax.plot(ph11.tiempo, ph6.promedio, label = 'ph 6')
ax.plot(ph11.tiempo, ph5.promedio, label = 'ph 5')

#ax.scatter(ph11.tiempo,pred)
ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.set_title('Durchschnitts - pH - Werte: 5 - 7')
ax.legend()




fig, ax = plt.subplots()
ax.plot(ph11.tiempo, ph7.promedio, label = 'ph 7')
ax.plot(ph11.tiempo, ph8.promedio, label = 'ph 8')
ax.plot(ph11.tiempo, ph9.promedio, label = 'ph 9')
ax.plot(ph11.tiempo, ph10.promedio, label = 'ph 10')
ax.plot(ph11.tiempo, ph11.promedio, label = 'ph 11')
#ax.scatter(ph11.tiempo,pred)
ax.set_xlabel('Zeit (s)')
ax.set_ylabel('Druck (kPa)')
ax.set_title('Durchschnitts - pH - Werte: 7 - 11')
ax.legend()



pendientes11a7 = [[coef7[0],'ph7'],[coef8[0],'ph8'],[coef9[0],'ph9'],[coef10[0],'ph10'],[coef[0],'ph11']]


pendientes7a5 = [[coef5[0],'ph5'],[coef6[0],'ph6'],[coef7[0],'ph7']]



phs7a5 = [i[1] for i in pendientes7a5]


pendientesDef7a5 = [i[0] for i in pendientes7a5]


phs11a7 = [i[1] for i in pendientes11a7]
pendientesDef11a7 = [i[0] for i in pendientes11a7]

finalPendiente = pendientesDef7a5 + pendientesDef11a7
finalPh = phs7a5+ phs11a7


promCoef11 = np.polyfit(ph11.tiempo, ph11.promedio,1)
promCoef10 = np.polyfit(ph11.tiempo, ph10.promedio,1)
promCoef9 = np.polyfit(ph11.tiempo, ph9.promedio,1)
promCoef8 = np.polyfit(ph11.tiempo, ph8.promedio,1)
promCoef7 = np.polyfit(ph11.tiempo, ph7.promedio,1)
promCoef6 = np.polyfit(ph11.tiempo, ph6.promedio,1)
promCoef5 = np.polyfit(ph11.tiempo, ph5.promedio,1)



pendientes11a7prom = [[promCoef7[0],'ph7'],[promCoef8[0],'ph8'],
                      [promCoef9[0],'ph9'],[promCoef10[0],'ph10'],
                      [promCoef11[0],'ph11']]


pendientes7a5prom = [[promCoef5[0],'ph5'],[promCoef6[0],'ph6'],
                     [promCoef7[0],'ph7']]


phs7a5prom = [i[1] for i in pendientes7a5prom]


pendientesDef7a5prom = [i[0] for i in pendientes7a5prom]
phs11a7prom = [i[1] for i in pendientes11a7prom]
pendientesDef11a7prom = [i[0] for i in pendientes11a7prom]

finalPendienteprom = pendientesDef7a5prom + pendientesDef11a7prom
finalPhprom = phs7a5prom+ phs11a7prom



fig, ax = plt.subplots()
ax.plot(finalPh, finalPendiente, label = 'Enzymaktivität')
#ax.scatter(ph11.tiempo,pred)
ax.set_xlabel('ph Wert')
ax.set_ylabel('Enzymaktivität')
ax.set_title('Enzymaktivität bei unterschieliche pH Werte (ausgewählt)')
ax.legend()

fig, ax = plt.subplots()
ax.plot(finalPhprom, finalPendienteprom, label = 'Durchschnittsenzymaktivität')
#ax.scatter(ph11.tiempo,pred)
ax.set_xlabel('ph Wert')
ax.set_ylabel('durchschnittliche Enzymaktivität')
ax.set_title('Enzymaktivität bei unterschieliche pH Werte (durchschnittlich)')
ax.legend()

